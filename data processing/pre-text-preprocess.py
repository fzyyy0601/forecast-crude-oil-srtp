'''
    py3.7
    writer :fzy
    date:2020/8/28
'''

import time
import re
from openpyxl import load_workbook
from nltk.corpus import stopwords
import nltk

from nltk.stem import WordNetLemmatizer
wd_lem=WordNetLemmatizer()

def info(filepath):
    wb=openpyxl.load_workbook(filepath)
    ws=wb.active
    print('{}行{}列'.format(ws.max_row,ws.max_column))
    '''
    col_range=ws['A:B']
    row_range=ws[1:3]
    for row in row_range:
        for cell in row:
            print(cell.value)
    '''

def readin_chunk():
#    data=pd.read_excel('d:\srtp\exp2.xlsx')
    workbook=load_workbook(u'd:\srtp\exp3.5.xlsx')
    rows=workbook.active.rows
    st=''
    print(time.asctime(time.localtime(time.time())))
    i=0
    for i in range(95000,10000):#按行读取 
        i=i+1
       # line=[col.value for col in row]
        c1=workbook.active.cell(row=i,column=1).value
        c2=workbook.active.cell(row=i,column=2).value
        st=st+str(c1)+str(c2)
        print(i)
    print(time.asctime(time.localtime(time.time())))
    return st

def process(txt):
    f=txt.lower()#小写
    f=re.sub(r'\d+','',f)#去除数字
    f=re.sub(r'[^\w]',' ',f)#去除标点符号
    f=wd_lem.lemmatize(f)
    text_list=nltk.word_tokenize(f)#分词
    stop=stopwords.words('english')#去除停用词
    
    text_list=[word for word in text_list if word not in stop and len(word)>2]
    print(time.asctime(time.localtime(time.time())))
    return(text_list)

def write_data(text,save_path):
    i=0
    with open(save_path,'a',encoding='utf-8') as f:
        for word in text:
            f.write(word)
            f.write(',')
         #   i+=1
         #   print(i)
    print(time.asctime(time.localtime(time.time())))


if __name__=='__main__':
   # print(time.asctime(time.localtime(time.time())))
    txt=readin_chunk()
    #print('--'*20)
    text_list=process(txt)
    save_path='d:\srtp\.results.txt'
    #print('--'*20)
    write_data(text_list,save_path)

